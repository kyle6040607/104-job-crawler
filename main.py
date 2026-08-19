"""104 人力銀行職缺爬蟲

用法：
    .venv\\Scripts\\python.exe main.py                       # 互動式，會逐項問你
    .venv\\Scripts\\python.exe main.py --keyword python --content 遠端 --exp 2,3 --edu 4 --limit 20
    .venv\\Scripts\\python.exe main.py --keyword 資料工程師 --title      # 標題必須含關鍵字
    .venv\\Scripts\\python.exe main.py --keyword 工程師 --exclude 結構,土木  # 排除不要的領域

重點摘要是直接從 104 原文抽條列做的，不需要任何 API 金鑰、不連外部 AI 服務。

輸出：終端機重點卡片 + Excel 檔（重點摘要、職缺網址、公司網址、公司簡介）
"""

import argparse
import datetime as dt
import re
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

for stream in (sys.stdout, sys.stderr, sys.stdin):
    try:
        stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")
SEARCH_API = "https://www.104.com.tw/jobs/search/api/jobs"
JOB_API = "https://www.104.com.tw/job/ajax/content/{}"
COMPANY_API = "https://www.104.com.tw/company/ajax/list"
DELAY = 0.6  # 每次請求間隔，別把對方打爆

# 104 的篩選代碼（實測確認）
EXP_OPTIONS = [("不拘", ""), ("1年以下", "1"), ("1~3年", "3"),
               ("3~5年", "5"), ("5~10年", "10"), ("10年以上", "99")]
EDU_OPTIONS = [("不拘", ""), ("高中以下", "1"), ("高中", "2"), ("專科", "3"),
               ("大學", "4"), ("碩士", "5"), ("博士", "6")]

session = requests.Session()
session.headers.update({"User-Agent": UA, "Accept": "application/json"})


def get_json(url, referer, params=None, retry=3):
    """帶 Referer 的 GET，104 沒有 Referer 會擋。失敗回 None。"""
    for i in range(retry):
        try:
            r = session.get(url, params=params, headers={"Referer": referer}, timeout=20)
            if r.status_code == 200 and "json" in r.headers.get("content-type", ""):
                return r.json()
        except requests.RequestException:
            pass
        time.sleep(1 + i)
    return None


def clean(text, limit=None):
    text = re.sub(r"\s+", " ", (text or "")).strip()
    return text[:limit] + "…" if limit and len(text) > limit else text


# ---------------------------------------------------------------- 選單


def pick_menu(title, options, preset=None):
    """讓使用者用編號複選，回傳 104 代碼字串（逗號分隔），不拘則回空字串。"""
    if preset is None:
        menu = "  ".join(f"{i}){name}" for i, (name, _) in enumerate(options))
        print(f"\n{title}")
        print(f"  {menu}")
        preset = input("  請輸入編號（可複選、逗號分隔，直接 Enter = 不拘）：").strip()
    idxs = [i.strip() for i in re.split(r"[,，\s]+", preset) if i.strip().isdigit()]
    codes = [options[int(i)][1] for i in idxs if 0 <= int(i) < len(options) and options[int(i)][1]]
    labels = [options[int(i)][0] for i in idxs if 0 <= int(i) < len(options) and options[int(i)][1]]
    return ",".join(codes), "、".join(labels) or "不拘"


# ---------------------------------------------------------------- 搜尋 / 抓取


def search_pages(keyword, jobexp="", edu="", exclude="", loose=False, max_pages=20):
    """依關鍵字逐頁搜尋職缺，一頁一頁 yield 出來。

    fz=0 是「精確比對」。104 預設 fz=8 會把關鍵字拆開模糊比對，
    「資料工程師」因此會撈到結構工程師、工地工程師這種完全無關的職缺。
    """
    seen = set()
    for page in range(1, max_pages + 1):
        params = {"keyword": keyword, "page": page, "pagesize": 20,
                  "order": 16, "asc": 0, "mode": "s", "jobsource": "index_s",
                  "fz": 8 if loose else 0}
        if jobexp:
            params["jobexp"] = jobexp
        if edu:
            params["edu"] = edu
        if exclude:
            params["excludeKeyword"] = exclude
        data = get_json(SEARCH_API, "https://www.104.com.tw/jobs/search/", params)
        batch = (data or {}).get("data") or []
        fresh = [j for j in batch if j.get("jobNo") not in seen]
        seen.update(j.get("jobNo") for j in fresh)
        if not fresh:
            return
        yield page, fresh
        time.sleep(DELAY)


def match_title(job, keyword):
    """標題必須沾到關鍵字（整串，或空白分隔的任一詞）才算數。"""
    title = (job.get("jobName") or "").lower()
    terms = [t for t in re.split(r"[\s,，/、]+", keyword.lower()) if t]
    return keyword.lower() in title or any(t in title for t in terms)


def match_content(job, words, require_all):
    """用搜尋結果的職缺描述先過濾，命中才去抓詳細頁，省請求數。"""
    if not words:
        return True, []
    haystack = " ".join([job.get("jobName", ""), job.get("description", ""),
                         job.get("custName", ""), job.get("coIndustryDesc", "")]).lower()
    hits = [w for w in words if w.lower() in haystack]
    return (len(hits) == len(words) if require_all else bool(hits)), hits


def job_detail(job_url):
    jid = job_url.rstrip("/").split("/")[-1].split("?")[0]
    data = get_json(JOB_API.format(jid), f"https://www.104.com.tw/job/{jid}")
    return (data or {}).get("data") or {}


_company_cache = {}


def company_profile(name, encoded_no):
    """用公司名稱搜尋公司庫，比對 encodedCustNo 取得公司簡介。"""
    if encoded_no in _company_cache:
        return _company_cache[encoded_no]
    data = get_json(COMPANY_API, "https://www.104.com.tw/company/search/",
                    {"keyword": name, "page": 1, "pageSize": 10})
    found = {}
    for c in (data or {}).get("data") or []:
        if c.get("encodedCustNo") == encoded_no:
            found = c
            break
    _company_cache[encoded_no] = found
    time.sleep(DELAY)
    return found


# ---------------------------------------------------------------- 摘要

NOISE = re.compile(r"(歡迎|意者請|請附|履歷|意者|本公司保留|符合資格|備註|上班地點|工作時間)")


MARKER = re.compile(r"^[\s]*(?:[•\-＊*※◆●▲★☆·・]|\(?\d{1,2}[\.\)、：:]|[（(]?[一二三四五六七八九十]{1,2}[)）、\.])")


def _bullets(text, limit=5):
    """把 104 那種條列式描述拆成重點行。優先取真正有編號/符號的條列。"""
    text = (text or "").replace("\r", "\n")
    marked, plain = [], []
    for raw in text.split("\n"):
        is_marked = bool(MARKER.match(raw))
        s = re.sub(r"^[\s•\-＊*※◆●▲★☆·・]+", "", raw).strip()
        s = re.sub(r"^\(?\d{1,2}[\.\)、：:]\s*", "", s)
        s = re.sub(r"^[（(]?[一二三四五六七八九十]{1,2}[)）、\.]\s*", "", s)
        if len(s) < 6 or NOISE.search(s):
            continue
        (marked if is_marked else plain).append(clean(s, 70))
    lines = marked if len(marked) >= 2 else (marked + plain)
    if not lines:  # 整段散文就退回切句子
        lines = [clean(s, 70) for s in re.split(r"[。；;\n]", text) if len(s.strip()) > 8]
    return lines[:limit]


def summarize_rule(detail, job):
    """規則式摘要：不需要 API 金鑰，直接從欄位抽重點。"""
    jd = detail.get("jobDetail") or {}
    cond = detail.get("condition") or {}
    desc = jd.get("jobDescription") or job.get("description", "")
    skills = [s.get("description", "") for s in cond.get("specialty") or []]
    parts = []
    duties = _bullets(desc, 4)
    if duties:
        parts.append("【工作內容】" + "；".join(duties))
    req = []
    if cond.get("workExp") and cond["workExp"] != "不拘":
        req.append(cond["workExp"])
    if cond.get("edu"):
        req.append(cond["edu"])
    if skills:
        req.append("技能：" + "、".join(skills))
    langs = ["/".join(filter(None, [l.get("language")])) for l in cond.get("language") or []]
    if langs:
        req.append("語言：" + "、".join(langs))
    if req:
        parts.append("【必備】" + "；".join(req))
    plus = _bullets(cond.get("other", ""), 2)
    if plus:
        parts.append("【加分】" + "；".join(plus))
    return "\n".join(parts)


# ---------------------------------------------------------------- 主流程


def collect(keyword, words, require_all, limit, jobexp, edu,
            exclude="", loose=False, title_only=False):
    rule = "（全部符合）" if require_all else "（任一符合）"
    mode = "模糊比對" if loose else "精確比對"
    print(f"\n[1/2] 搜尋「{keyword}」（{mode}"
          f"{'、標題須含關鍵字' if title_only else ''}"
          f"{'、排除：' + exclude if exclude else ''}）")
    print(f"      內容關鍵字：{'、'.join(words) + rule if words else '（無，不過濾）'}")

    picked, scanned, dropped = [], 0, 0
    for page, batch in search_pages(keyword, jobexp, edu, exclude, loose):
        scanned += len(batch)
        for job in batch:
            if title_only and not match_title(job, keyword):
                dropped += 1
                continue
            ok, hits = match_content(job, words, require_all)
            if ok:
                picked.append((job, hits))
        tail = f"，標題不符剔除 {dropped} 筆" if title_only else ""
        print(f"      第 {page} 頁：掃描 {scanned} 筆，符合 {len(picked)} 筆{tail}")
        if len(picked) >= limit:
            break
    picked = picked[:limit]
    if not picked:
        return []

    print(f"\n[2/2] 抓取 {len(picked)} 筆職缺詳細內容與公司簡介…")
    rows = []
    for i, (job, hits) in enumerate(picked, 1):
        link = job.get("link") or {}
        job_url = link.get("job", "")
        cust_url = link.get("cust", "")
        detail = job_detail(job_url)
        time.sleep(DELAY)

        jd = detail.get("jobDetail") or {}
        cond = detail.get("condition") or {}
        cust_id = cust_url.rstrip("/").split("/")[-1].split("?")[0]
        comp = company_profile(job.get("custName", ""), cust_id)

        rows.append({
            "職稱": job.get("jobName", ""),
            "公司名稱": job.get("custName", ""),
            "產業": job.get("coIndustryDesc", ""),
            "工作地點": job.get("jobAddrNoDesc", "") + " " + (jd.get("addressDetail") or job.get("jobAddress", "")),
            "薪資": jd.get("salary") or "未提供",
            "工作經歷": cond.get("workExp", ""),
            "學歷": cond.get("edu", ""),
            "技能/工具": "、".join(s.get("description", "") for s in cond.get("specialty") or []),
            "更新日期": job.get("appearDate", ""),
            "命中內容關鍵字": "、".join(hits),
            "重點摘要": summarize_rule(detail, job),
            "福利": "、".join((detail.get("welfare") or {}).get("tag") or []),
            "職缺網址": job_url,
            "公司網址": cust_url,
            "公司資本額": comp.get("capitalDesc", ""),
            "公司員工數": comp.get("employeeCountDesc", "") or (f"{job.get('employeeCount')}人" if job.get("employeeCount") else ""),
            "公司簡介": clean(comp.get("profile", "")),
            "職缺內容全文": clean(jd.get("jobDescription") or job.get("description", "")),
        })
        print(f"      ({i}/{len(picked)}) {rows[-1]['公司名稱']} — {rows[-1]['職稱']}")
    return rows


def print_summary(rows):
    print("\n" + "=" * 78)
    for i, r in enumerate(rows, 1):
        print(f"\n{i}. {r['職稱']}")
        print(f"   公司：{r['公司名稱']}（{r['產業']}）")
        print(f"   地點：{r['工作地點'].strip()}　薪資：{r['薪資']}")
        print(f"   條件：{r['工作經歷']} / {r['學歷']}" + (f" / {r['技能/工具']}" if r["技能/工具"] else ""))
        if r["命中內容關鍵字"]:
            print(f"   命中：{r['命中內容關鍵字']}")
        for line in r["重點摘要"].split("\n"):
            print(f"   ▸ {line}")
        print(f"   公司簡介：{clean(r['公司簡介'], 130) or '（該公司未提供）'}")
        print(f"   職缺：{r['職缺網址']}")
        print(f"   公司：{r['公司網址']}")
    print("\n" + "=" * 78)


def save_excel(rows, keyword):
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[\\/:*?"<>|\s]', "_", keyword) or "104"
    path = f"104_{safe}_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "職缺"
    headers = list(rows[0].keys())
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="2F5597")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for r in rows:
        ws.append([r[h] for h in headers])

    widths = {"職稱": 34, "公司名稱": 24, "產業": 20, "工作地點": 30, "薪資": 16,
              "工作經歷": 12, "學歷": 12, "技能/工具": 22, "更新日期": 12,
              "命中內容關鍵字": 16, "重點摘要": 60, "福利": 36,
              "職缺網址": 38, "公司網址": 38, "公司資本額": 16, "公司員工數": 14,
              "公司簡介": 60, "職缺內容全文": 70}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 18)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, h in enumerate(headers):
            if h.endswith("網址") and row[idx].value:
                row[idx].hyperlink = row[idx].value
                row[idx].font = Font(color="0563C1", underline="single")
        row[0].parent.row_dimensions[row[0].row].height = 90
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(path)
    return path


def main():
    p = argparse.ArgumentParser(description="104 職缺爬蟲")
    p.add_argument("--keyword", help="職缺關鍵字")
    p.add_argument("--content", nargs="?", const="", default=None,
                   help="內容關鍵字，逗號分隔；只寫 --content 不給值代表不過濾")
    p.add_argument("--exp", help="工作經歷選單編號：0不拘 1(1年以下) 2(1~3年) 3(3~5年) 4(5~10年) 5(10年以上)")
    p.add_argument("--edu", help="學歷選單編號：0不拘 1高中以下 2高中 3專科 4大學 5碩士 6博士")
    p.add_argument("--limit", type=int, help="最多要幾筆職缺")
    p.add_argument("--exclude", nargs="?", const="", default=None,
                   help="排除關鍵字，逗號分隔（例：結構,土木,工地）；只寫 --exclude 代表不排除")
    p.add_argument("--title", action="store_true", help="只保留標題含關鍵字的職缺（最嚴格）")
    p.add_argument("--loose", action="store_true", help="用 104 預設的模糊比對（會撈到很多不相干的）")
    p.add_argument("--any", action="store_true", help="內容關鍵字改成「任一符合」（預設全部符合）")
    args = p.parse_args()

    keyword = args.keyword or input("職缺關鍵字（例：python 工程師）：").strip()
    if not keyword:
        print("沒有輸入關鍵字，結束。")
        return
    content = args.content if args.content is not None else input("內容關鍵字（逗號分隔，可留空）：").strip()
    words = [w.strip() for w in re.split(r"[,，]", content) if w.strip()]
    exclude = args.exclude if args.exclude is not None else input("排除關鍵字（逗號分隔，可留空，例：結構,土木）：").strip()
    exclude = ",".join(w.strip() for w in re.split(r"[,，]", exclude) if w.strip())

    jobexp, exp_label = pick_menu("工作經歷（104 篩選）", EXP_OPTIONS, args.exp)
    edu, edu_label = pick_menu("學歷（104 篩選）", EDU_OPTIONS, args.edu)
    print(f"\n  已選：工作經歷 = {exp_label}／學歷 = {edu_label}")

    limit = args.limit or int(input("\n要抓幾筆？（預設 20）：").strip() or 20)

    rows = collect(keyword, words, not args.any, limit, jobexp, edu,
                   exclude, args.loose, args.title)
    if not rows:
        print("\n沒有符合條件的職缺，放寬經歷／學歷或換個關鍵字試試。")
        return
    print_summary(rows)
    path = save_excel(rows, keyword)
    print(f"\n已輸出 Excel：{path}（共 {len(rows)} 筆）")


if __name__ == "__main__":
    main()
